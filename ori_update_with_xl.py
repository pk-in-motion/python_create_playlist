"""
https://developer.dailymotion.com/video-upload/how-to-upload-videos-dailymotion-excel/
"""

# Import the dailymotion sdk
import importlib
import dailymotion
from datetime import datetime

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


# dailymotion credentials
_API_KEY = "77f772cfc5172ded7b66"
_API_SECRET = "b25de6f8c4c87a9ddc5554eec994c4a9e1b60808"
_USERNAME = "pk.in.motion@protonmail.ch"
_PASSWORD = "1L0veDM_2019"

# Loading the workbook
_FILENAME = "Videos.xlsx"
_XLS = load_workbook(_FILENAME)

# Change your sheet name if necessary
_SHEET = _XLS.get_sheet_by_name("Sheet1")

# Adding some info returned by the API on the sheet

_DM_UPDTIME = _SHEET.cell(
    row=1, column=_SHEET.max_column + 1, value="Updated Time"
)


# Instanciating the sdk
_DM = dailymotion.Dailymotion()

try:

    # Providing the authentication information
    _DM.set_grant_type(
        "password",
        api_key=_API_KEY,
        api_secret=_API_SECRET,
        scope=["manage_videos"],
        info={"username": _USERNAME, "password": _PASSWORD},
    )

    # Looping into the rows and skiping the headers
    for row in _SHEET[2 : _SHEET.max_row]:

        # Uploading the file on dailymotion servers - commented if no upload
        # url = _DM.upload(row[1].value)

        # Filling the information about the video
        parameters = {
            #"url": url,
            "title": row[0].value,
            "tags": row[2].value,
            "channel": row[3].value,
            "description": row[4].value,
            "private": row[5].value,
        }

        # Sending the information to create the video on dailymotion
        # ori => result = _DM.post("/me/videos?fields=id,url", parameters)
        result = _DM.post("/video/" + row[1].value + "?fields=id,updated_time", parameters)

        unixTime = (result['updated_time'])
        normalTime = datetime.utcfromtimestamp(unixTime).strftime('%Y-%m-%d %H:%M:%S')


        # write the xid returned by the API
        _DM_UPDTIME = _SHEET.cell(
            row=_DM_UPDTIME.row + 1, column=_DM_UPDTIME.column, value=normalTime
        )


    print("Upload from excel file to dailymotion finished")
    _XLS.save(_FILENAME)

except Exception as e:
    print("An error occured: %s" % str(e))
